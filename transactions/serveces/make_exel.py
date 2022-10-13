import openpyxl
from openpyxl.styles import Font, PatternFill

green_RGB = "C1E3C9"
red_RGB = "F3C5CB"
blue_RGB = "007AFD"


def make_color_transaction_xlsx(transaction, row):
    if transaction.money_value > 0:
        for cell in row:
            cell.fill = PatternFill(fgColor=green_RGB, fill_type="solid")
    else:
        for cell in row:
            cell.fill = PatternFill(fgColor=red_RGB, fill_type="solid")


def completion_transactions_xlsx(list_table, count_transactions: int, transactions):
    rows = list_table.iter_rows(min_row=2, max_col=5, max_row=count_transactions + 1)
    for row, transactions_address in zip(rows, range(count_transactions)):
        list_table.row_dimensions[row[0].row].height = 25
        transaction = transactions[transactions_address]
        row[0].value = transaction.data_time.replace(tzinfo=None)
        row[1].value = transaction.money_value
        row[2].value = transaction.accounts.title
        row[3].value = transaction.transactions_type.category
        row[4].value = transaction.comment

        make_color_transaction_xlsx(transaction, row)


def make_xlsx_layout(list_table, transactions):
    list_table.title = "transactions history"
    list_table.column_dimensions['A'].width = 30
    list_table.column_dimensions['B'].width = 30
    list_table.column_dimensions['C'].width = 20
    list_table.column_dimensions['D'].width = 20
    list_table.column_dimensions['E'].width = 20
    list_table['A1'] = "Дата добавления"
    list_table['B1'] = "Денежная сумма"
    list_table['C1'] = "Счёт"
    list_table['D1'] = "Категория"
    list_table['E1'] = "Примечания"
    for cell in list_table['1']:
        cell.font = Font(size=15, color="FFFFFF", bold=True, italic=True)
        cell.fill = PatternFill(fgColor=blue_RGB, fill_type="solid")


def make_xlsx_file_in_response(response, transactions):
    wb = openpyxl.Workbook()
    list_table = wb.active
    make_xlsx_layout(list_table, transactions)
    completion_transactions_xlsx(list_table, transactions.count(), transactions)
    wb.save(response)
    wb.close()
    return response
